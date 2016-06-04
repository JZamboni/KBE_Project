from openpyxl import Workbook
from openpyxl import load_workbook

class ExcelImporter:

    def __init__(self, filePath):
        self.filePath=filePath
        self.excelFile = load_workbook(filePath)
        self.sheet=self.excelFile['Sheet1']



    def readData(self):

        wb = Workbook()
        importValue = self.sheet['D5'].value
        myFuselageLength=fuselageLength.fuselageLength()
        fuselageLength.fuselageLength.setValue(myFuselageLength, value=importValue)

        return fuselageLength.fuselageLength.getValue(myFuselageLength)